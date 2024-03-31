from playwright.sync_api import sync_playwright
from bs4 import BeautifulSoup
from datetime import datetime
import openpyxl as xl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, LineChart
from email.message import EmailMessage
import ssl
import smtplib


"""
The first part of the script uses playwright to reach the online employee KPI data.
GreenTech uses Tamago to manage this data. Naturally, I have removed the username and password from the
input fields for this public version as well as the names of all individuals involved in this project.
I create lists of each KPI for the current week, then lists for the year to date.
These lists are put in the 'latest_data_file.xlsx' to be read from and used in the 'KPI.xlsx'.
The 'KPI.xlsx' contains 2 sheets per staff. One with KPI data per week, year to date and targets.
The other 'main' sheet has all KPI data visualised in bar and line charts showing: 
past week, past 4 weeks, and past year for each KPI vs individual targets.
Ratios for ICI vs CV, and CV vs Placed are also given by quarter and year to date.
"""

current_datetime = datetime.now()
current_date = current_datetime.date()
current_year = current_datetime.year
current_month = current_datetime.month
current_day = current_datetime.day
date_object = datetime(current_year, current_month, current_day)
current_week_of_year = date_object.isocalendar()[1]
month_end_week = [4, 8, 12, 16, 20, 24, 28, 32, 36, 40, 44, 48, 52]
annual_ratio_row = 43

if current_week_of_year <= 12:
    quarter = 1
if 13 <= current_week_of_year <= 24:
    quarter = 2
if 25 <= current_week_of_year <= 36:
    quarter = 3
if 37 <= current_week_of_year <= 48:
    quarter = 4

if quarter == 1:
    row_of_quarter = 39
if quarter == 2:
    row_of_quarter = 40
if quarter == 3:
    row_of_quarter = 41
if quarter == 4:
    row_of_quarter = 42


def main():
    with sync_playwright() as p:
        page_url = 'https://titan.tamago-db.com/job/hiring'
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(page_url, timeout=30000)
        page.is_visible('#login-btn')
        page.click('#login-btn')
        page.is_visible('#password')
        page.fill('#username', '**********')
        page.fill('#password', '**********')
        page.click('#_submit')
        page.wait_for_timeout(1000)
        page.is_visible('#menu-container')
        try:
            page.click('#main-menu > li.last.dropdown')
        except TimeoutError:
            page.click('#main-menu > li.last.dropdown.open > a')
        page.is_visible('#main-menu > li.last.dropdown.open > ul > li:nth-child(3) > a > span')
        page.click('#main-menu > li.last.dropdown.open > ul > li:nth-child(3) > a > span')
        page.is_visible('#grid_performance_report > div.grid_body.table-responsive > table > thead >'
                        ' tr.grid-row-filters > th:nth-child(3) > span > span > span > div > button')
        page.click('#grid_performance_report > div.grid_body.table-responsive > table > thead > tr.grid-row-filters > '
                   'th:nth-child(3) > span > span > span > div > button')
        page.wait_for_timeout(1000)
        page.check('#grid_performance_report > div.grid_body.table-responsive > table > thead >'
                   ' tr.grid-row-filters > th:nth-child(3) > span > span > span > div > ul > '
                   'li:nth-child(2) > a > label > input[type=checkbox]')
        page.click('#grid_performance_report > div.grid_body.table-responsive > table > thead > tr.grid-row-filters > '
                   'th.last-column')
        page.click('#reports-datepicker')
        page.wait_for_timeout(1000)
        page.click('body > div.daterangepicker.dropdown-menu.opensright > div.ranges > ul > li:nth-child(3)')
        html = page.inner_html('#grid_performance_report > div.grid_body.table-responsive > table > tbody')
        soup = BeautifulSoup(html, 'html.parser')
        date = []
        name = []
        ici = []
        cvs = []
        ccm1 = []
        placed = []
        cl = []
        f2f = []
        pch = []
        year_date = []
        year_name = []
        year_ici = []
        year_cvs = []
        year_ccm1 = []
        year_placed = []
        year_cl = []
        year_f2f = []
        year_pch = []

        for row in soup.select('tr'):
            columns = row.select('td')
            if not columns:
                continue
            date.append(columns[0].text.strip())
            name.append(columns[1].text.strip())
            ici.append(columns[3].text.strip())
            cvs.append(columns[5].text.strip())
            ccm1.append(columns[6].text.strip())
            placed.append(columns[7].text.strip())
            cl.append(columns[8].text.strip())
        page.click('#reports-datepicker')
        page.click('body > div.daterangepicker.dropdown-menu.opensright > div.ranges > ul > li:nth-child(9)')
        html = page.inner_html('#grid_performance_report > div.grid_body.table-responsive > table > tbody')
        soup = BeautifulSoup(html, 'html.parser')
        for row in soup.select('tr'):
            columns = row.select('td')
            if not columns:
                continue
            year_date.append(columns[0].text.strip())
            year_name.append(columns[1].text.strip())
            year_ici.append(columns[3].text.strip())
            year_cvs.append(columns[5].text.strip())
            year_ccm1.append(columns[6].text.strip())
            year_placed.append(columns[7].text.strip())
            year_cl.append(columns[8].text.strip())
        try:
            page.click('#main-menu > li.last.dropdown')
        except TimeoutError:
            page.click('#main-menu > li.last.dropdown.open > a')
        page.is_visible('#main-menu > li.active.last.dropdown.open > ul > li:nth-child(4) > a > span')
        page.click('#main-menu > li.active.last.dropdown.open > ul > li:nth-child(4) > a > span')
        page.is_visible('#grid_events_report > div.grid_body.table-responsive > table > thead > tr.grid-row-filters > '
                        'th:nth-child(3) > span > span > span > div > button')
        page.click('#grid_events_report > div.grid_body.table-responsive > table > thead > tr.grid-row-filters > '
                   'th:nth-child(3) > span > span > span > div > button')
        page.check('#grid_events_report > div.grid_body.table-responsive > table > thead > tr.grid-row-filters > '
                   'th:nth-child(3) > span > span > span > div > ul > li:nth-child(2) > a > '
                   'label > input[type=checkbox]')
        page.click('#grid_events_report > div.grid_body.table-responsive > table > thead > tr.grid-row-filters > '
                   'th.last-column > button')
        page.click('#reports-datepicker')
        page.wait_for_timeout(1000)
        page.click('body > div.daterangepicker.dropdown-menu.opensright > div.ranges > ul > li:nth-child(3)')

        html = page.inner_html('#grid_events_report > div.grid_body.table-responsive > table > tbody')
        soup = BeautifulSoup(html, 'html.parser')
        for row in soup.select('tr'):
            columns = row.select('td')
            if not columns:
                continue
            f2f.append(columns[15].text.strip())
            pch.append(columns[17].text.strip())

        staff_dict_this_week = {'date': date, 'name': name, 'ici': ici, 'cvs': cvs,
                                'ccm1': ccm1, 'cl': cl, 'f2f': f2f, 'pch': pch, 'placed': placed}

        page.click('#reports-datepicker')
        page.click('body > div.daterangepicker.dropdown-menu.opensright > div.ranges > ul > li:nth-child(9)')
        html = page.inner_html('#grid_events_report > div.grid_body.table-responsive > table > tbody')
        soup = BeautifulSoup(html, 'html.parser')
        for row in soup.select('tr'):
            columns = row.select('td')
            if not columns:
                continue
            year_f2f.append(columns[15].text.strip())
            year_pch.append(columns[17].text.strip())

        staff_dict_this_year = {'date': year_date, 'name': year_name, 'ici': year_ici, 'cvs': year_cvs,
                                'ccm1': year_ccm1, 'cl': year_cl, 'f2f': year_f2f, 'pch': year_pch,
                                'placed': year_placed}
        browser.close()

    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = 'This Week'
    sheet2 = workbook.create_sheet(title='This Year')

    def write_data_to_sheet(sheet, data):
        header = list(data.keys())
        sheet.append(header)
        for row_data in zip(*data.values()):
            sheet.append(row_data)

    write_data_to_sheet(sheet1, staff_dict_this_week)
    write_data_to_sheet(sheet2, staff_dict_this_year)
    workbook.save('latest_data_file.xlsx')

    latest_data_file = xl.load_workbook('latest_data_file.xlsx')
    data_file = xl.load_workbook('KPI.xlsx')
    this_week_sheet = latest_data_file['This Week']
    this_year_sheet = latest_data_file['This Year']
    if len(data_file.sheetnames) < 2:
        ws1 = data_file.create_sheet('Sheet1', 0)
        ws1.title = "Team Overview"
        ws2 = data_file.create_sheet("Sheet2", 1)
        ws2.title = "Mr Black"
        ws3 = data_file.create_sheet("Sheet3", 2)
        ws3.title = "Mr White"
        ws4 = data_file.create_sheet("Sheet4", 3)
        ws4.title = "Mr Brown"
        ws5 = data_file.create_sheet("Sheet5", 4)
        ws5.title = "Mr Pink"
        ws6 = data_file.create_sheet("Sheet6", 5)
        ws6.title = "Mr Blue"
        ws7 = data_file.create_sheet("Sheet7", 6)
        ws7.title = "Beatrix"
        ws8 = data_file.create_sheet('Sheet8', 7)
        ws8.title = "Vince"
        ws9 = data_file.create_sheet("Sheet9", 8)
        ws9.title = "Mr Black1"
        ws10 = data_file.create_sheet("Sheet10", 9)
        ws10.title = "Mr White1"
        ws11 = data_file.create_sheet("Sheet11", 10)
        ws11.title = "Mr Brown1"
        ws12 = data_file.create_sheet("Sheet12", 11)
        ws12.title = "Mr Pink1"
        ws13 = data_file.create_sheet("Sheet13", 12)
        ws13.title = "Mr Blue1"
        ws14 = data_file.create_sheet("Sheet14", 13)
        ws14.title = "Beatrix1"
        ws15 = data_file.create_sheet("Sheet15", 14)
        ws15.title = "Vince1"

    def add_date_data(sheet_name):
        sheet = sheet_name
        cell = sheet.cell(current_week_of_year + 1, 1)
        cell.value = str(current_date)

    def add_ici_data(sheet_name, target, done):
        sheet = sheet_name
        cell = sheet.cell(1, 2)
        cell.value = "ICI"
        cell = sheet.cell(1, 3)
        cell.value = "ICI Target"
        cell = sheet.cell(current_week_of_year + 1, 2)
        cell.value = int(done.value)
        cell = sheet.cell(current_week_of_year + 1, 3)
        cell.value = target

    def add_pch_data(sheet_name, target, done):
        sheet = sheet_name
        if sheet == brown_sheet:
            cell = sheet.cell(1, 2)
            cell.value = "PCH"
            cell = sheet.cell(1, 3)
            cell.value = "PCH Target"
            cell = sheet.cell(current_week_of_year + 1, 2)
            cell.value = int(done.value)
            cell = sheet.cell(current_week_of_year + 1, 3)
            cell.value = target

        elif sheet == beatrix_sheet or sheet == g:
            cell = sheet.cell(1, 9)
            cell.value = "PCH"
            cell = sheet.cell(1, 10)
            cell.value = "PCH Target"
            cell = sheet.cell(current_week_of_year + 1, 9)
            cell.value = int(done.value)
            cell = sheet.cell(current_week_of_year + 1, 10)
            cell.value = target
        else:
            cell = sheet.cell(1, 8)
            cell.value = "PCH"
            cell = sheet.cell(1, 9)
            cell.value = "PCH Target"
            cell = sheet.cell(current_week_of_year + 1, 8)
            cell.value = int(done.value)
            cell = sheet.cell(current_week_of_year + 1, 9)
            cell.value = target

    def add_cvs_data(sheet_name, target, done):
        sheet = sheet_name
        cell = sheet.cell(1, 4)
        cell.value = "CV"
        cell = sheet.cell(1, 5)
        cell.value = "CV Target"
        cell = sheet.cell(current_week_of_year + 1, 4)
        cell.value = int(done.value)
        cell = sheet.cell(current_week_of_year + 1, 5)
        cell.value = target

    def add_ccm1_data(sheet_name, target, done):
        sheet = sheet_name
        cell = sheet.cell(1, 6)
        cell.value = "CCM1"
        cell = sheet.cell(1, 7)
        cell.value = "CCM1 Target"
        cell = sheet.cell(current_week_of_year + 1, 6)
        cell.value = float(done.value)
        cell = sheet.cell(current_week_of_year + 1, 7)
        cell.value = target

    def add_placed_data(sheet_name, done):
        sheet = sheet_name
        cell = sheet.cell(1, 8)
        cell.value = 'Placed'
        cell = sheet.cell(current_week_of_year + 1, 8)
        cell.value = int(done.value)

    def add_cl_data(sheet_name, target, done):
        sheet = sheet_name
        cell = sheet.cell(1, 9)
        cell.value = "CL"
        cell = sheet.cell(1, 10)
        cell.value = "CL Target"
        cell = sheet.cell(current_week_of_year + 1, 9)
        cell.value = int(done.value)
        cell = sheet.cell(current_week_of_year + 1, 10)
        cell.value = target

    def add_f2f_data(sheet_name, target, done):
        sheet = sheet_name
        cell = sheet.cell(1, 11)
        cell.value = "F2F"
        cell = sheet.cell(1, 12)
        cell.value = "F2F Target"
        cell = sheet.cell(current_week_of_year + 1, 11)
        cell.value = int(done.value)
        cell = sheet.cell(current_week_of_year + 1, 12)
        cell.value = target

    def add_data_year_to_date(sheet_name, ici, ici_target, cv, cv_target, ccm1, ccm1_target, placed, cl, cl_target, f2f,
                              f2f_target, pch, pch_target):
        sheet = sheet_name
        cell = sheet.cell(1, 13)
        cell.value = "Year ICI"
        cell = sheet.cell(1, 14)
        cell.value = "Year ICI Target"
        cell = sheet.cell(1, 15)
        cell.value = "Year CV"
        cell = sheet.cell(1, 16)
        cell.value = "Year CV Target"
        cell = sheet.cell(1, 17)
        cell.value = "Year CCM1"
        cell = sheet.cell(1, 18)
        cell.value = "Year CCM1 Target"
        cell = sheet.cell(1, 19)
        cell.value = "Year Placed"
        cell = sheet.cell(1, 20)
        cell.value = "Year CL"
        cell = sheet.cell(1, 21)
        cell.value = "Year CL Target"
        cell = sheet.cell(1, 22)
        cell.value = "Year F2F"
        cell = sheet.cell(1, 23)
        cell.value = "Year F2F Target"
        cell = sheet.cell(1, 24)
        cell.value = "Year PCH"
        cell = sheet.cell(1, 25)
        cell.value = "Year PCH Target"
        cell = sheet.cell(current_week_of_year + 1, 13)
        cell.value = int(ici.value)
        cell = sheet.cell(current_week_of_year + 1, 14)
        cell.value = current_week_of_year * ici_target
        cell = sheet.cell(current_week_of_year + 1, 15)
        cell.value = int(cv.value)
        cell = sheet.cell(current_week_of_year + 1, 16)
        cell.value = current_week_of_year * cv_target
        cell = sheet.cell(current_week_of_year + 1, 17)
        cell.value = int(ccm1.value)
        cell = sheet.cell(current_week_of_year + 1, 18)
        cell.value = current_week_of_year * ccm1_target
        cell = sheet.cell(current_week_of_year + 1, 19)
        cell.value = int(placed.value)
        cell = sheet.cell(current_week_of_year + 1, 20)
        cell.value = int(cl.value)
        cell = sheet.cell(current_week_of_year + 1, 21)
        cell.value = current_week_of_year * cl_target
        cell = sheet.cell(current_week_of_year + 1, 22)
        cell.value = int(f2f.value)
        cell = sheet.cell(current_week_of_year + 1, 23)
        cell.value = current_week_of_year * f2f_target
        cell = sheet.cell(current_week_of_year + 1, 24)
        cell.value = int(pch.value)
        cell = sheet.cell(current_week_of_year + 1, 25)
        cell.value = current_week_of_year * pch_target

    def update_ratio_table(graph_sheet, sheet, cv_done_col, ccm1_done_col, placed_col):
        ratio_table_cv_title_cell = graph_sheet.cell(38, 3)
        ratio_table_cv_title_cell.value = "CV"
        ratio_table_ccm1_title_cell = graph_sheet.cell(38, 4)
        ratio_table_ccm1_title_cell.value = "CCM1"
        ratio_table_placed_cell = graph_sheet.cell(38, 5)
        ratio_table_placed_cell.value = "Placed"
        ratio_table_cv_to_ccm1_title_cell = graph_sheet.cell(38, 6)
        ratio_table_cv_to_ccm1_title_cell.value = "CV-CCM1"
        ratio_table_ccm1_to_placed_title_cell = graph_sheet.cell(38, 7)
        ratio_table_ccm1_to_placed_title_cell.value = "CCM1-Placed"
        ratio_table_annual_title_cell = graph_sheet.cell(annual_ratio_row, 2)
        ratio_table_annual_title_cell.value = "Annual"
        annual_ratio_table_cv_cell = graph_sheet.cell(annual_ratio_row, 3)
        annual_ratio_table_cv_cell.value = sheet.cell(current_week_of_year + 1, cv_done_col).value
        annual_ratio_table_ccm1_cell = graph_sheet.cell(annual_ratio_row, 4)
        annual_ratio_table_ccm1_cell.value = sheet.cell(current_week_of_year + 1, ccm1_done_col).value
        annual_ratio_table_placed_cell = graph_sheet.cell(annual_ratio_row, 5)
        annual_ratio_table_placed_cell.value = sheet.cell(current_week_of_year + 1, placed_col).value
        annual_ratio_table_cv_ccm1_ratio_cell = graph_sheet.cell(annual_ratio_row, 6)
        if annual_ratio_table_ccm1_cell.value >= 1:
            annual_ratio_table_cv_ccm1_ratio_cell.value = annual_ratio_table_cv_cell.value / annual_ratio_table_ccm1_cell.value
        else:
            annual_ratio_table_cv_ccm1_ratio_cell.value = "NA"
        annual_ratio_table_ccm1_placed_ratio_cell = graph_sheet.cell(annual_ratio_row, 7)
        if annual_ratio_table_placed_cell.value >= 1:
            annual_ratio_table_ccm1_placed_ratio_cell.value = annual_ratio_table_ccm1_cell.value / annual_ratio_table_placed_cell.value
        else:
            annual_ratio_table_ccm1_placed_ratio_cell.value = "NA"

        if quarter == 1:
            ratio_table_quarter_1_title_cell = graph_sheet.cell(39, 2)
            ratio_table_quarter_1_title_cell.value = "Q1"
            ratio_table_cv_cell = graph_sheet.cell(row_of_quarter, 3)
            ratio_table_cv_cell.value = sheet.cell(current_week_of_year + 1, cv_done_col).value
            ratio_table_ccm1_cell = graph_sheet.cell(row_of_quarter, 4)
            ratio_table_ccm1_cell.value = sheet.cell(current_week_of_year + 1, ccm1_done_col).value
            ratio_table_placed_cell = graph_sheet.cell(row_of_quarter, 5)
            ratio_table_placed_cell.value = sheet.cell(current_week_of_year + 1, placed_col).value
            ratio_table_cv_ccm1_ratio_cell = graph_sheet.cell(row_of_quarter, 6)
            # If the value is 0, the division can't take place. Therefore, the value of NA is shown.
            if ratio_table_ccm1_cell.value >= 1:
                ratio_table_cv_ccm1_ratio_cell.value = ratio_table_cv_cell.value / ratio_table_ccm1_cell.value
            else:
                ratio_table_cv_ccm1_ratio_cell.value = "NA"
            ratio_table_ccm1_placed_ratio_cell = graph_sheet.cell(row_of_quarter, 7)
            if ratio_table_placed_cell.value >= 1:
                ratio_table_ccm1_placed_ratio_cell.value = ratio_table_ccm1_cell.value / ratio_table_placed_cell.value
            else:
                ratio_table_ccm1_placed_ratio_cell.value = "NA"
        if quarter == 2:
            ratio_table_quarter_2_title_cell = graph_sheet.cell(40, 2)
            ratio_table_quarter_2_title_cell.value = "Q2"
            ratio_table_cv_cell = graph_sheet.cell(row_of_quarter, 3)
            ratio_table_cv_cell.value = (sheet.cell(current_week_of_year + 1, cv_done_col).value -
                                         graph_sheet.cell(39, 3).value)
            ratio_table_ccm1_cell = graph_sheet.cell(row_of_quarter, 4)
            ratio_table_ccm1_cell.value = (sheet.cell(current_week_of_year + 1, ccm1_done_col).value -
                                           graph_sheet.cell(39, 4).value)
            ratio_table_placed_cell = graph_sheet.cell(row_of_quarter, 5)
            ratio_table_placed_cell.value = sheet.cell(current_week_of_year + 1, placed_col).value - graph_sheet.cell(39, 5).value
            ratio_table_cv_ccm1_ratio_cell = graph_sheet.cell(row_of_quarter, 6)
            if ratio_table_ccm1_cell.value >= 1:
                ratio_table_cv_ccm1_ratio_cell.value = ratio_table_cv_cell.value / ratio_table_ccm1_cell.value
            else:
                ratio_table_cv_ccm1_ratio_cell.value = "NA"
            ratio_table_ccm1_placed_ratio_cell = graph_sheet.cell(row_of_quarter, 7)
            if ratio_table_placed_cell.value >= 1:
                ratio_table_ccm1_placed_ratio_cell.value = ratio_table_ccm1_cell.value / ratio_table_placed_cell.value
            else:
                ratio_table_ccm1_placed_ratio_cell.value = "NA"

        if quarter == 3:
            ratio_table_quarter_2_title_cell = sheet.cell(41, 2)
            ratio_table_quarter_2_title_cell.value = "Q3"
            ratio_table_cv_cell = graph_sheet.cell(row_of_quarter, 3)
            ratio_table_cv_cell.value = (sheet.cell(current_week_of_year + 1, cv_done_col).value -
                                         graph_sheet.cell(39, 3).value - graph_sheet.cell(40, 3).value)
            ratio_table_ccm1_cell = graph_sheet.cell(row_of_quarter, 4)
            ratio_table_ccm1_cell.value = (sheet.cell(current_week_of_year + 1, ccm1_done_col).value -
                                           graph_sheet.cell(39, 4).value - graph_sheet.cell(40, 4).value)
            ratio_table_placed_cell = graph_sheet.cell(row_of_quarter, 5)
            ratio_table_placed_cell.value = (sheet.cell(current_week_of_year + 1, placed_col).value -
                                             graph_sheet.cell(39, 5).value - graph_sheet.cell(40, 5).value)
            ratio_table_cv_ccm1_ratio_cell = graph_sheet.cell(row_of_quarter, 6)
            if ratio_table_ccm1_cell.value >= 1:
                ratio_table_cv_ccm1_ratio_cell.value = ratio_table_cv_cell.value / ratio_table_ccm1_cell.value
            else:
                ratio_table_cv_ccm1_ratio_cell.value = "NA"
            ratio_table_ccm1_placed_ratio_cell = graph_sheet.cell(row_of_quarter, 7)
            if ratio_table_placed_cell.value >= 1:
                ratio_table_ccm1_placed_ratio_cell.value = ratio_table_ccm1_cell.value / ratio_table_placed_cell.value
            else:
                ratio_table_ccm1_placed_ratio_cell.value = "NA"

        if quarter == 4:
            ratio_table_quarter_2_title_cell = graph_sheet.cell(42, 2)
            ratio_table_quarter_2_title_cell.value = "Q4"
            ratio_table_cv_cell = graph_sheet.cell(row_of_quarter, 3)
            ratio_table_cv_cell.value = (sheet.cell(current_week_of_year + 1, cv_done_col).value -
                                         graph_sheet.cell.value(39, 3)
                                         - graph_sheet.cell(40, 3).value - graph_sheet.cell(41, 3).value)
            ratio_table_ccm1_cell = graph_sheet.cell(row_of_quarter, 4)
            ratio_table_ccm1_cell.value = (sheet.cell(current_week_of_year + 1, ccm1_done_col).value -
                                           graph_sheet.cell(39, 4).value - graph_sheet.cell(40, 3).value -
                                           graph_sheet.cell(41, 3).value)
            ratio_table_placed_cell = graph_sheet.cell(row_of_quarter, 5)
            ratio_table_placed_cell.value = (sheet.cell(current_week_of_year + 1, placed_col).value -
                                             graph_sheet.cell(39, 5).value - graph_sheet.cell(40, 3).value -
                                             graph_sheet.cell(41, 3).value)
            ratio_table_cv_ccm1_ratio_cell = graph_sheet.cell(row_of_quarter, 6)
            if ratio_table_ccm1_cell.value >= 1:
                ratio_table_cv_ccm1_ratio_cell.value = ratio_table_cv_cell.value / ratio_table_ccm1_cell.value
            else:
                ratio_table_cv_ccm1_ratio_cell.value = "NA"
            ratio_table_ccm1_placed_ratio_cell = graph_sheet.cell(row_of_quarter, 7)
            if ratio_table_placed_cell.value >= 1:
                ratio_table_ccm1_placed_ratio_cell.value = ratio_table_ccm1_cell.value / ratio_table_placed_cell.value
            else:
                ratio_table_ccm1_placed_ratio_cell.value = "NA"

    def add_bar_chart_weekly(sheet_name, graph_sheet_name, column_1, column_2, chart_title, x_title, y_title, placement,
                             ):
        sheet = sheet_name
        graph_sheet = graph_sheet_name
        done_values = Reference(sheet,
                                min_row=current_week_of_year + 1,
                                max_row=current_week_of_year + 1,
                                min_col=column_1,
                                max_col=column_1)
        target_values = Reference(sheet,
                                  min_row=current_week_of_year + 1,
                                  max_row=current_week_of_year + 1,
                                  min_col=column_2,
                                  max_col=column_2)
        chart = BarChart()
        chart.title = chart_title
        chart.y_axis.title = y_title
        chart.x_axis.title = x_title
        chart.add_data(done_values)
        chart.add_data(target_values)
        graph_sheet.add_chart(chart, placement)

    def add_line_chart_annual(sheet_name, graph_sheet_name, column_1, column_2, chart_title, x_title, y_title,
                              placement):
        sheet = sheet_name
        graph_sheet = graph_sheet_name
        year_done_values = Reference(sheet,
                                     min_row=1,
                                     max_row=49,
                                     min_col=column_1,
                                     max_col=column_1)
        year_target_values = Reference(sheet,
                                       min_row=1,
                                       max_row=49,
                                       min_col=column_2,
                                       max_col=column_2)
        chart = LineChart()
        chart.title = chart_title
        chart.y_axis.title = y_title
        chart.x_axis.title = x_title
        chart.add_data(year_done_values, titles_from_data=True)
        chart.add_data(year_target_values, titles_from_data=True)
        graph_sheet.add_chart(chart, placement)

    def add_line_chart_trending_monthly(sheet_name, graph_sheet_name, column_1, column_2, chart_title, x_title, y_title,
                                        placement):
        sheet = sheet_name
        graph_sheet = graph_sheet_name
        past_month_values = Reference(sheet,
                                      min_row=current_week_of_year - 2,
                                      max_row=current_week_of_year + 1,
                                      min_col=column_1,
                                      max_col=column_1)
        past_month_target_values = Reference(sheet,
                                             min_row=current_week_of_year - 2,
                                             max_row=current_week_of_year + 1,
                                             min_col=column_2,
                                             max_col=column_2)
        chart = LineChart()
        chart.title = chart_title
        chart.y_axis.title = y_title
        chart.x_axis.title = x_title
        chart.add_data(past_month_values)
        chart.add_data(past_month_target_values)
        graph_sheet.add_chart(chart, placement)

    def remove_all_charts(sheet):
        for chart in sheet._charts:
            sheet._charts.remove(chart)


    black_sheet = data_file['Mr Black1']
    black_graph_sheet = data_file['Mr Black']
    black_ici_done = this_week_sheet.cell(2, 3)
    black_cvs_done = this_week_sheet.cell(2, 4)
    black_ccm1_done = this_week_sheet.cell(2, 5)
    black_cl_done = this_week_sheet.cell(2, 6)
    black_f2f_done = this_week_sheet.cell(2, 7)
    black_placed_done = this_week_sheet.cell(2, 9)
    black_ici_done_year = this_year_sheet.cell(2, 3)
    black_cvs_done_year = this_year_sheet.cell(2, 4)
    black_ccm1_done_year = this_year_sheet.cell(2, 5)
    black_cl_done_year = this_year_sheet.cell(2, 6)
    black_f2f_done_year = this_year_sheet.cell(2, 7)
    black_pch_done_year = this_year_sheet.cell(2, 8)
    black_placed_done_year = this_year_sheet.cell(2, 9)

    add_date_data(black_sheet)
    add_ici_data(black_sheet, 1, black_ici_done)
    add_cvs_data(black_sheet, 3, black_cvs_done)
    add_ccm1_data(black_sheet, 1.5, black_ccm1_done)
    add_placed_data(black_sheet, black_placed_done)
    add_cl_data(black_sheet, 3, black_cl_done)
    add_f2f_data(black_sheet, 1, black_f2f_done)
    add_data_year_to_date(black_sheet, black_ici_done_year, 1, black_cvs_done_year, 3,
                          black_ccm1_done_year, 1.5, black_placed_done_year, black_cl_done_year,
                          3, black_f2f_done_year, 1, black_pch_done_year, 0)

    remove_all_charts(black_graph_sheet)

    add_bar_chart_weekly(black_sheet, black_graph_sheet, 2, 3,
                         "ICI Done vs Target", f"Week of {current_date}",
                         "number", "b2")
    add_bar_chart_weekly(black_sheet, black_graph_sheet, 4, 5,
                         "CV Done vs Target", f"Week of {current_date}",
                         "number", "j2")
    add_bar_chart_weekly(black_sheet, black_graph_sheet, 6, 7,
                         "CCM1 Done vs Target", f"Week of {current_date}",
                         "number", "r2")
    add_bar_chart_weekly(black_sheet, black_graph_sheet, 9, 10,
                         "CL Done vs Target", f"Week of {current_date}",
                         "number", "z2")
    add_bar_chart_weekly(black_sheet, black_graph_sheet, 11, 12,
                         "F2F Done vs Target", f"Week of {current_date}",
                         "number", "AH2")

    add_line_chart_trending_monthly(black_sheet, black_graph_sheet, 2, 3,
                                    "Past month ICI",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "b14")
    add_line_chart_trending_monthly(black_sheet, black_graph_sheet, 4, 5,
                                    "Past month CV",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "j14")
    add_line_chart_trending_monthly(black_sheet, black_graph_sheet, 6, 7,
                                    "Past month CCM1",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "r14")
    add_line_chart_trending_monthly(black_sheet, black_graph_sheet, 9, 10,
                                    "Past month CL",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "z14")
    add_line_chart_trending_monthly(black_sheet, black_graph_sheet, 11, 12,
                                    "Past month F2F",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "AH14")

    add_line_chart_annual(black_sheet, black_graph_sheet, 13, 14, "Annual ICI",
                          f"{current_year} by week", "number", "b26")
    add_line_chart_annual(black_sheet, black_graph_sheet, 15, 16, "Annual CV",
                          f"{current_year} by week", "number", "j26")
    add_line_chart_annual(black_sheet, black_graph_sheet, 17, 18, "Annual CCM1",
                          f"{current_year} by week", "number", "r26")
    add_line_chart_annual(black_sheet, black_graph_sheet, 20, 21, "Annual CL",
                          f"{current_year} by week", "number", "z26")
    add_line_chart_annual(black_sheet, black_graph_sheet, 22, 23, "Annual F2F",
                          f"{current_year} by week", "number", "AH26")

    update_ratio_table(black_graph_sheet, black_sheet, 15, 17, 19)

    white_sheet = data_file['Mr White1']
    white_graph_sheet = data_file['Mr White']
    white_ici_sheet = this_week_sheet.cell(4, 3)
    white_cvs_done = this_week_sheet.cell(4, 4)
    white_ccm1_done = this_week_sheet.cell(4, 5)
    white_cl_done = this_week_sheet.cell(4, 6)
    white_f2f_done = this_week_sheet.cell(4, 7)
    white_placed_done = this_week_sheet.cell(4, 9)
    white_ici_done_year = this_year_sheet.cell(4, 3)
    white_cvs_done_year = this_year_sheet.cell(4, 4)
    white_ccm1_done_year = this_year_sheet.cell(4, 5)
    white_cl_done_year = this_year_sheet.cell(4, 6)
    white_f2f_done_year = this_year_sheet.cell(4, 7)
    white_pch_done_year = this_year_sheet.cell(4, 8)
    white_placed_done_year = this_year_sheet.cell(4, 9)

    add_date_data(white_sheet)
    add_ici_data(white_sheet, 3, white_ici_sheet)
    add_cvs_data(white_sheet, 6, white_cvs_done)
    add_ccm1_data(white_sheet, 2, white_ccm1_done)
    add_placed_data(white_sheet, white_placed_done)
    add_cl_data(white_sheet, 1, white_cl_done)
    add_f2f_data(white_sheet, 1, white_f2f_done)
    add_data_year_to_date(white_sheet, white_ici_done_year, 3, white_cvs_done_year, 6,
                          white_ccm1_done_year, 2, white_placed_done_year, white_cl_done_year,
                          1, white_f2f_done_year, 1, white_pch_done_year, 0)

    remove_all_charts(white_graph_sheet)

    add_bar_chart_weekly(white_sheet, white_graph_sheet, 2, 3,
                         "ICI Done vs Target", f"Week of {current_date}",
                         "number", "b2")
    add_bar_chart_weekly(white_sheet, white_graph_sheet, 4, 5,
                         "CV Done vs Target", f"Week of {current_date}",
                         "number", "j2")
    add_bar_chart_weekly(white_sheet, white_graph_sheet, 6, 7,
                         "CCM1 Done vs Target", f"Week of {current_date}",
                         "number", "r2")
    add_bar_chart_weekly(white_sheet, white_graph_sheet, 8, 9,
                         "CL Done vs Target", f"Week of {current_date}",
                         "number", "z2")
    add_bar_chart_weekly(white_sheet, white_graph_sheet, 10, 11,
                         "F2F Done vs Target", f"Week of {current_date}",
                         "number", "AH2")

    add_line_chart_trending_monthly(white_sheet, white_graph_sheet, 2, 3,
                                    "Past month ICI",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "b14")
    add_line_chart_trending_monthly(white_sheet, white_graph_sheet, 4, 5,
                                    "Past month CV",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "j14")
    add_line_chart_trending_monthly(white_sheet, white_graph_sheet, 6, 7,
                                    "Past month CCM1",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "r14")
    add_line_chart_trending_monthly(white_sheet, white_graph_sheet, 9, 10,
                                    "Past month CL",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "z14")
    add_line_chart_trending_monthly(white_sheet, white_graph_sheet, 11, 12,
                                    "Past month F2F",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "AH14")

    add_line_chart_annual(white_sheet, white_graph_sheet, 13, 14, "Annual ICI",
                          f"{current_year} by week", "number", "b26")
    add_line_chart_annual(white_sheet, white_graph_sheet, 15, 16, "Annual CV",
                          f"{current_year} by week", "number", "j26")
    add_line_chart_annual(white_sheet, white_graph_sheet, 17, 18, "Annual CCM1",
                          f"{current_year} by week", "number", "r26")
    add_line_chart_annual(white_sheet, white_graph_sheet, 20, 21, "Annual CL",
                          f"{current_year} by week", "number", "z26")
    add_line_chart_annual(white_sheet, white_graph_sheet, 22, 23, "Annual F2F",
                          f"{current_year} by week", "number", "AH26")
    update_ratio_table(white_graph_sheet, white_sheet, 15, 17, 19)

    brown_sheet = data_file['Mr Brown1']
    brown_graph_sheet = data_file['Mr Brown']
    brown_pch_done = this_week_sheet.cell(5, 8)
    brown_placed_done = this_week_sheet.cell(5, 9)
    brown_ici_done_year = this_year_sheet.cell(5, 3)
    brown_cvs_done_year = this_year_sheet.cell(5, 4)
    brown_ccm1_done_year = this_year_sheet.cell(5, 5)
    brown_cl_done_year = this_year_sheet.cell(5, 6)
    brown_f2f_done_year = this_year_sheet.cell(5, 7)
    brown_pch_done_year = this_year_sheet.cell(5, 8)
    brown_placed_done_year = this_year_sheet.cell(5, 9)

    add_date_data(brown_sheet)
    add_pch_data(brown_sheet, 6, brown_pch_done)
    add_data_year_to_date(brown_sheet, brown_ici_done_year, 0, brown_cvs_done_year, 0,
                          brown_ccm1_done_year, 0, brown_placed_done_year, brown_cl_done_year,
                          0, brown_f2f_done_year, 0, brown_pch_done_year, 6)

    remove_all_charts(brown_graph_sheet)

    add_bar_chart_weekly(brown_sheet, brown_graph_sheet, 2, 3,
                         "PCH Done vs Target", f"Week of {current_date}",
                         "number", "b2")

    add_line_chart_trending_monthly(brown_sheet, brown_graph_sheet, 2, 3,
                                    "Past month PCH",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "b14")

    add_line_chart_annual(brown_sheet, brown_graph_sheet, 24, 25, "Annual PCH",
                          f"{current_year} by week", "number", "b26")

    pink_sheet = data_file['Mr Pink1']
    pink_graph_sheet = data_file['Mr Pink']
    pink_ici_done = this_week_sheet.cell(6, 3)
    pink_cvs_done = this_week_sheet.cell(6, 4)
    pink_ccm1_done = this_week_sheet.cell(6, 5)
    pink_cl_done = this_week_sheet.cell(6, 6)
    pink_f2f_done = this_week_sheet.cell(6, 7)
    pink_placed_done = this_week_sheet.cell(6, 9)
    pink_ici_done_year = this_year_sheet.cell(6, 3)
    pink_cvs_done_year = this_year_sheet.cell(6, 4)
    pink_ccm1_done_year = this_year_sheet.cell(6, 5)
    pink_cl_done_year = this_year_sheet.cell(6, 6)
    pink_f2f_done_year = this_year_sheet.cell(6, 7)
    pink_pch_done_year = this_year_sheet.cell(6, 8)
    pink_placed_done_year = this_year_sheet.cell(6, 9)

    add_date_data(pink_sheet)
    add_ici_data(pink_sheet, 3, pink_ici_done)
    add_cvs_data(pink_sheet, 6, pink_cvs_done)
    add_ccm1_data(pink_sheet, 2, pink_ccm1_done)
    add_placed_data(pink_sheet, pink_placed_done)
    add_cl_data(pink_sheet, 1, pink_cl_done)
    add_f2f_data(pink_sheet, 1, pink_f2f_done)
    add_data_year_to_date(pink_sheet, pink_ici_done_year, 3, pink_cvs_done_year, 6,
                          pink_ccm1_done_year, 2, pink_placed_done_year, pink_cl_done_year,
                          1, pink_f2f_done_year, 1, pink_pch_done_year, 6)

    remove_all_charts(pink_graph_sheet)

    add_bar_chart_weekly(pink_sheet, pink_graph_sheet, 2, 3,
                         "ICI Done vs Target", f"Week of {current_date}",
                         "number", "b2")
    add_bar_chart_weekly(pink_sheet, pink_graph_sheet, 4, 5,
                         "CV Done vs Target", f"Week of {current_date}",
                         "number", "j2")
    add_bar_chart_weekly(pink_sheet, pink_graph_sheet, 6, 7,
                         "CCM1 Done vs Target", f"Week of {current_date}",
                         "number", "r2")
    add_bar_chart_weekly(pink_sheet, pink_graph_sheet, 9, 10,
                         "CL Done vs Target", f"Week of {current_date}",
                         "number", "z2")
    add_bar_chart_weekly(pink_sheet, pink_graph_sheet, 11, 12,
                         "F2F Done vs Target", f"Week of {current_date}",
                         "number", "AH2")

    add_line_chart_trending_monthly(pink_sheet, pink_graph_sheet, 2, 3,
                                    "Past month ICI",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "b14")
    add_line_chart_trending_monthly(pink_sheet, pink_graph_sheet, 4, 5,
                                    "Past month CV",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "j14")
    add_line_chart_trending_monthly(pink_sheet, pink_graph_sheet, 6, 7,
                                    "Past month CCM1",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "r14")
    add_line_chart_trending_monthly(pink_sheet, pink_graph_sheet, 9, 10,
                                    "Past month CL",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "z14")
    add_line_chart_trending_monthly(pink_sheet, pink_graph_sheet, 11, 12,
                                    "Past month F2F",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "AH14")

    add_line_chart_annual(pink_sheet, pink_graph_sheet, 13, 14, "Annual ICI",
                          f"{current_year} by week", "number", "b26")
    add_line_chart_annual(pink_sheet, pink_graph_sheet, 15, 16, "Annual CV",
                          f"{current_year} by week", "number", "j26")
    add_line_chart_annual(pink_sheet, pink_graph_sheet, 17, 18, "Annual CCM1",
                          f"{current_year} by week", "number", "r26")
    add_line_chart_annual(pink_sheet, pink_graph_sheet, 20, 21, "Annual CL",
                          f"{current_year} by week", "number", "z26")
    add_line_chart_annual(pink_sheet, pink_graph_sheet, 22, 23, "Annual F2F",
                          f"{current_year} by week", "number", "AH26")
    update_ratio_table(pink_graph_sheet, pink_sheet, 15, 17, 19)

    blue_sheet = data_file['Mr Blue1']
    blue_graph_sheet = data_file['Mr Blue']
    blue_ici_done = this_week_sheet.cell(7, 3)
    blue_cvs_done = this_week_sheet.cell(7, 4)
    blue_ccm1_done = this_week_sheet.cell(7, 5)
    blue_cl_done = this_week_sheet.cell(7, 6)
    blue_f2f_done = this_week_sheet.cell(7, 7)
    blue_placed_done = this_week_sheet.cell(7, 9)
    blue_ici_done_year = this_year_sheet.cell(7, 3)
    blue_cvs_done_year = this_year_sheet.cell(7, 4)
    blue_ccm1_done_year = this_year_sheet.cell(7, 5)
    blue_placed_done_year = this_year_sheet.cell(7, 6)
    blue_cl_done_year = this_year_sheet.cell(7, 6)
    blue_f2f_done_year = this_year_sheet.cell(7, 7)
    blue_pch_done_year = this_year_sheet.cell(7, 8)

    add_date_data(blue_sheet)
    add_ici_data(blue_sheet, 3, blue_ici_done)
    add_cvs_data(blue_sheet, 7, blue_cvs_done)
    add_ccm1_data(blue_sheet, 2.2, blue_ccm1_done)
    add_placed_data(blue_sheet, blue_placed_done)
    add_cl_data(blue_sheet, 1, blue_cl_done)
    add_f2f_data(blue_sheet, 1, blue_f2f_done)
    add_data_year_to_date(blue_sheet, blue_ici_done_year, 3, blue_cvs_done_year, 7,
                          blue_ccm1_done_year, 2.2, blue_placed_done_year, blue_cl_done_year,
                          1, blue_f2f_done_year, 1, blue_pch_done_year, 0)

    remove_all_charts(blue_graph_sheet)

    add_bar_chart_weekly(blue_sheet, blue_graph_sheet, 2, 3,
                         "ICI Done vs Target", f"Week of {current_date}",
                         "number", "b2")
    add_bar_chart_weekly(blue_sheet, blue_graph_sheet, 4, 5,
                         "CV Done vs Target", f"Week of {current_date}",
                         "number", "j2")
    add_bar_chart_weekly(blue_sheet, blue_graph_sheet, 6, 7,
                         "CCM1 Done vs Target", f"Week of {current_date}",
                         "number", "r2")
    add_bar_chart_weekly(blue_sheet, blue_graph_sheet, 9, 10,
                         "CL Done vs Target", f"Week of {current_date}",
                         "number", "z2")
    add_bar_chart_weekly(blue_sheet, blue_graph_sheet, 11, 12,
                         "F2F Done vs Target", f"Week of {current_date}",
                         "number", "AH2")

    add_line_chart_trending_monthly(blue_sheet, blue_graph_sheet, 2, 3,
                                    "Past month ICI",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "b14")
    add_line_chart_trending_monthly(blue_sheet, blue_graph_sheet, 4, 5,
                                    "Past month CV",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "j14")
    add_line_chart_trending_monthly(blue_sheet, blue_graph_sheet, 6, 7,
                                    "Past month CCM1",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "r14")
    add_line_chart_trending_monthly(blue_sheet, blue_graph_sheet, 9, 10,
                                    "Past month CL",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "z14")
    add_line_chart_trending_monthly(blue_sheet, blue_graph_sheet, 11, 12,
                                    "Past month F2F",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "AH14")

    add_line_chart_annual(blue_sheet, blue_graph_sheet, 13, 14, "Annual ICI",
                          f"{current_year} by week", "number", "b26")
    add_line_chart_annual(blue_sheet, blue_graph_sheet, 15, 16, "Annual CV",
                          f"{current_year} by week", "number", "j26")
    add_line_chart_annual(blue_sheet, blue_graph_sheet, 17, 18, "Annual CCM1",
                          f"{current_year} by week", "number", "r26")
    add_line_chart_annual(blue_sheet, blue_graph_sheet, 20, 21, "Annual CL",
                          f"{current_year} by week", "number", "z26")
    add_line_chart_annual(blue_sheet, blue_graph_sheet, 22, 23, "Annual F2F",
                          f"{current_year} by week", "number", "AH26")
    update_ratio_table(blue_graph_sheet, blue_sheet, 15, 17, 19)

    beatrix_sheet = data_file['Beatrix1']
    beatrix_graph_sheet = data_file['Beatrix']
    beatrix_ici_done = this_week_sheet.cell(8, 3)
    beatrix_cvs_done = this_week_sheet.cell(8, 4)
    beatrix_ccm1_done = this_week_sheet.cell(8, 5)
    beatrix_pch_done = this_week_sheet.cell(8, 8)
    beatrix_placed_done = this_week_sheet.cell(8, 9)
    beatrix_ici_done_year = this_year_sheet.cell(8, 3)
    beatrix_cvs_done_year = this_year_sheet.cell(8, 4)
    beatrix_ccm1_done_year = this_year_sheet.cell(8, 5)
    beatrix_placed_done_year = this_year_sheet.cell(8, 9)
    beatrix_cl_done_year = this_year_sheet.cell(8, 6)
    beatrix_f2f_done_year = this_year_sheet.cell(8, 7)
    beatrix_pch_done_year = this_year_sheet.cell(8, 8)

    add_date_data(beatrix_sheet)
    add_pch_data(beatrix_sheet, 5, beatrix_pch_done)
    add_ici_data(beatrix_sheet, 3, beatrix_ici_done)
    add_cvs_data(beatrix_sheet, 3, beatrix_cvs_done)
    add_ccm1_data(beatrix_sheet, 1, beatrix_ccm1_done)
    add_placed_data(beatrix_sheet, beatrix_placed_done)
    add_data_year_to_date(beatrix_sheet, beatrix_ici_done_year, 3, beatrix_cvs_done_year, 3,
                          beatrix_ccm1_done_year, 1, beatrix_placed_done_year, beatrix_cl_done_year,
                          0, beatrix_f2f_done_year, 0, beatrix_pch_done_year, 5)

    remove_all_charts(beatrix_graph_sheet)

    add_bar_chart_weekly(beatrix_sheet, beatrix_graph_sheet, 2, 3,
                         "ICI Done vs Target", f"Week of {current_date}",
                         "number", "b2")
    add_bar_chart_weekly(beatrix_sheet, beatrix_graph_sheet, 4, 5,
                         "CV Done vs Target", f"Week of {current_date}",
                         "number", "j2")
    add_bar_chart_weekly(beatrix_sheet, beatrix_graph_sheet, 6, 7,
                         "CCM1 Done vs Target", f"Week of {current_date}",
                         "number", "r2")
    add_bar_chart_weekly(beatrix_sheet, beatrix_graph_sheet, 8, 9,
                         "PCH Done vs Target", f"Week of {current_date}",
                         "number", "r2")

    add_line_chart_trending_monthly(beatrix_sheet, beatrix_graph_sheet, 2, 3,
                                    "Past month ICI",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "b14")
    add_line_chart_trending_monthly(beatrix_sheet, beatrix_graph_sheet, 4, 5,
                                    "Past month CV",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "j14")
    add_line_chart_trending_monthly(beatrix_sheet, beatrix_graph_sheet, 9, 10,
                                    "Past month PCH",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "r14")

    add_line_chart_annual(beatrix_sheet, beatrix_graph_sheet, 13, 14, "Annual ICI",
                          f"{current_year} by week", "number", "b26")
    add_line_chart_annual(beatrix_sheet, beatrix_graph_sheet, 15, 16, "Annual CV",
                          f"{current_year} by week", "number", "j26")
    add_line_chart_annual(beatrix_sheet, beatrix_graph_sheet, 24, 25, "Annual PCH",
                          f"{current_year} by week", "number", "r26")
    update_ratio_table(beatrix_graph_sheet, beatrix_sheet, 15, 17, 19)

    g = data_file['miss g']
    g_graph_sheet = data_file['miss g']
    g_ici_done = this_week_sheet.cell(9, 3)
    g_cvs_done = this_week_sheet.cell(9, 4)
    g_ccm1_done = this_week_sheet.cell(9, 5)
    g_pch_done = this_week_sheet.cell(9, 8)
    g_placed_done = this_week_sheet.cell(9, 9)
    g_ici_done_year = this_year_sheet.cell(9, 3)
    g_cvs_done_year = this_year_sheet.cell(9, 4)
    g_ccm1_done_year = this_year_sheet.cell(9, 5)
    g_cl_done_year = this_year_sheet.cell(9, 6)
    g_f2f_done_year = this_year_sheet.cell(9, 7)
    g_pch_done_year = this_year_sheet.cell(9, 8)
    g_placed_done_year = this_year_sheet.cell(9, 9)

    add_date_data(g)
    add_pch_data(g, 5, g_pch_done)
    add_ici_data(g, 3, g_ici_done)
    add_cvs_data(g, 3, g_cvs_done)
    add_ccm1_data(g, 1, g_ccm1_done)
    add_placed_data(g, g_placed_done)
    add_data_year_to_date(g, g_ici_done_year, 3, g_cvs_done_year, 3,
                          g_ccm1_done_year, 1, g_placed_done_year, g_cl_done_year,
                          0, g_f2f_done_year, 0, g_pch_done_year, 5)

    remove_all_charts(g_graph_sheet)

    add_bar_chart_weekly(g, g_graph_sheet, 2, 3,
                         "ICI Done vs Target", f"Week of {current_date}",
                         "number", "b2")
    add_bar_chart_weekly(g, g_graph_sheet, 4, 5,
                         "CV Done vs Target", f"Week of {current_date}",
                         "number", "j2")
    add_bar_chart_weekly(g, g_graph_sheet, 6, 7,
                         "CCM1 Done vs Target", f"Week of {current_date}",
                         "number", "r2")
    add_bar_chart_weekly(g, g_graph_sheet, 8, 9,
                         "PCH Done vs Target", f"Week of {current_date}",
                         "number", "r2")

    add_line_chart_trending_monthly(g, g_graph_sheet, 2, 3,
                                    "Past month ICI",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "b14")
    add_line_chart_trending_monthly(g, g_graph_sheet, 4, 5,
                                    "Past month CV",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "j14")
    add_line_chart_trending_monthly(g, g_graph_sheet, 9, 10,
                                    "Past month PCH",
                                    f"Week {current_week_of_year - 3} to week {current_week_of_year}",
                                    "number", "r14")

    add_line_chart_annual(g, g_graph_sheet, 13, 14, "Annual ICI",
                          f"{current_year} by week", "number", "b26")
    add_line_chart_annual(g, g_graph_sheet, 15, 16, "Annual CV",
                          f"{current_year} by week", "number", "j26")
    add_line_chart_annual(g, g_graph_sheet, 24, 25, "Annual PCH",
                          f"{current_year} by week", "number", "r26")
    update_ratio_table(g_graph_sheet, g, 15, 17, 19)

    latest_data_file.save("latest_data_file.xlsx")
    data_file.save("KPI.xlsx")

    # writing and sending email
    email_sender = '******************@gmail.com'
    email_password = '**********************'
    email_receivers = ['******************@gmail.com', '**************@**************.jp']
    subject = "KPI Update"
    body = f"""
    <html>
        <body>
            <img src="https://greentech.titanconsulting.jp/wp-content/uploads/2023/03/Titan_GreenTech_logo-removebg-preview.png" style="max-width:25%; max-height:25%;">
            <br>
            <h1>KPI Update</h1>
        </body>
    </html>
    """

    # Files to attach
    data_file = "KPI.xlsx"

    em = EmailMessage()
    em['From'] = email_sender
    em['To'] = email_receivers
    em['Subject'] = subject
    em.set_content(body, subtype='html')

    with open(data_file, "rb") as f:
        em.add_attachment(f.read(), filename="KPI.xlsx", maintype="application",
                          subtype="octet-stream")

    context = ssl.create_default_context()

    with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
        smtp.login(email_sender, email_password)
        smtp.sendmail(email_sender, email_receivers, em.as_string())


if __name__ == '__main__':
    main()
